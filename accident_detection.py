import argparse
import cv2
import cvzone
import math
import time
import os
import numpy as np
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from datetime import datetime
import glob
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
from PIL import Image, ImageTk
import threading
from pathlib import Path


class AccidentDetectionGUI:
    def __init__(self, root):

        self.root = root
        self.root.title("Accident Detection System")
        self.root.geometry("1200x800")
        self.root.configure(bg="#f0f0f0")

        self.primary_color = "#2c3e50"
        self.accent_color = "#3498db"
        self.text_color = "#ecf0f1"
        self.bg_color = "#f0f0f0"
        self.success_color = "#2ecc71"
        self.warning_color = "#e74c3c"

        self.model = None
        self.model_loaded = False
        self.classNames = ['moderate', 'severe']
        self.confidence_threshold = 0.7
        self.detection_time_threshold = 2.0
        self.batch_size = 1

        # Video/image processing variables
        self.file_paths = []
        self.current_file_index = 0
        self.is_video_playing = False
        self.cap = None
        self.current_frame = None
        self.accident_count = 0
        self.total_accidents = 0
        self.global_accident_count = 0  # New: Global counter that persists across sessions
        self.accident_locations = []

        # Processing thread
        self.processing_thread = None
        self.stop_processing = False

        # Create GUI elements
        self.create_menu()
        self.create_main_layout()
        self.create_status_bar()

        # Load model on startup
        self.load_model()

        # Load global accident count from file if exists
        self.load_global_count()

    def create_menu(self):
        """Create the application menu"""
        menu_bar = tk.Menu(self.root)

        # File menu
        file_menu = tk.Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Load Model", command=self.load_model)
        file_menu.add_command(label="Select Files", command=self.select_files)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.root.quit)
        menu_bar.add_cascade(label="File", menu=file_menu)

        # View menu
        view_menu = tk.Menu(menu_bar, tearoff=0)
        view_menu.add_command(label="View Log", command=self.view_log)
        menu_bar.add_cascade(label="View", menu=view_menu)

        # Help menu
        help_menu = tk.Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="About", command=self.show_about)
        menu_bar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menu_bar)

    def create_main_layout(self):
        """Create the main application layout"""
        # Main frame
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Left panel - controls
        control_frame = ttk.LabelFrame(main_frame, text="Controls", padding=10)
        control_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))

        # File selection
        file_frame = ttk.Frame(control_frame)
        file_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(file_frame, text="Select Files", command=self.select_files).pack(fill=tk.X)

        # Files listbox
        self.files_frame = ttk.LabelFrame(control_frame, text="Selected Files")
        self.files_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.files_listbox = tk.Listbox(self.files_frame, height=10, selectmode=tk.SINGLE)
        self.files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.files_listbox.bind('<<ListboxSelect>>', self.on_file_select)

        files_scrollbar = ttk.Scrollbar(self.files_frame, orient=tk.VERTICAL, command=self.files_listbox.yview)
        files_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.files_listbox.config(yscrollcommand=files_scrollbar.set)

        # Parameters frame
        params_frame = ttk.LabelFrame(control_frame, text="Detection Parameters")
        params_frame.pack(fill=tk.X, pady=(0, 10))

        # Confidence threshold
        ttk.Label(params_frame, text="Confidence Threshold:").pack(anchor=tk.W)
        self.confidence_var = tk.DoubleVar(value=self.confidence_threshold)
        confidence_scale = ttk.Scale(params_frame, from_=0.1, to=1.0, variable=self.confidence_var,
                                     orient=tk.HORIZONTAL, length=200)
        confidence_scale.pack(fill=tk.X)
        ttk.Label(params_frame, textvariable=self.confidence_var).pack(anchor=tk.E)

        # Persistence threshold
        ttk.Label(params_frame, text="Persistence Threshold (s):").pack(anchor=tk.W)
        self.persistence_var = tk.DoubleVar(value=self.detection_time_threshold)
        persistence_scale = ttk.Scale(params_frame, from_=0.5, to=5.0, variable=self.persistence_var,
                                      orient=tk.HORIZONTAL, length=200)
        persistence_scale.pack(fill=tk.X)
        ttk.Label(params_frame, textvariable=self.persistence_var).pack(anchor=tk.E)

        # Action buttons
        actions_frame = ttk.Frame(control_frame)
        actions_frame.pack(fill=tk.X, pady=(0, 10))

        self.process_btn = ttk.Button(actions_frame, text="Process Selected", command=self.process_selected)
        self.process_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.stop_btn = ttk.Button(actions_frame, text="Stop", command=self.stop_processing_cmd, state=tk.DISABLED)
        self.stop_btn.pack(side=tk.LEFT)

        # Results frame
        results_frame = ttk.LabelFrame(control_frame, text="Results")
        results_frame.pack(fill=tk.X)

        ttk.Label(results_frame, text="Accidents in current file:").grid(row=0, column=0, sticky=tk.W)
        self.current_accidents_var = tk.StringVar(value="0")
        ttk.Label(results_frame, textvariable=self.current_accidents_var, font=('Arial', 12, 'bold')).grid(row=0,
                                                                                                           column=1,
                                                                                                           sticky=tk.E)

        ttk.Label(results_frame, text="Total accidents:").grid(row=1, column=0, sticky=tk.W)
        self.total_accidents_var = tk.StringVar(value="0")
        ttk.Label(results_frame, textvariable=self.total_accidents_var, font=('Arial', 12, 'bold')).grid(row=1,
                                                                                                         column=1,
                                                                                                         sticky=tk.E)

        # Right panel - display
        display_frame = ttk.LabelFrame(main_frame, text="Detection View", padding=10)
        display_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

        # Canvas for displaying images/videos
        self.canvas = tk.Canvas(display_frame, bg="black")
        self.canvas.pack(fill=tk.BOTH, expand=True)

        # Media controls frame
        media_controls = ttk.Frame(display_frame)
        media_controls.pack(fill=tk.X, pady=(10, 0))

        self.play_btn = ttk.Button(media_controls, text="▶ Play", command=self.toggle_play, state=tk.DISABLED)
        self.play_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.prev_btn = ttk.Button(media_controls, text="⏮ Previous", command=self.previous_file, state=tk.DISABLED)
        self.prev_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.next_btn = ttk.Button(media_controls, text="Next ⏭", command=self.next_file, state=tk.DISABLED)
        self.next_btn.pack(side=tk.LEFT)

    def create_status_bar(self):
        """Create the status bar at the bottom of the application"""
        self.status_var = tk.StringVar(value="Ready")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def load_model(self):
        """Load the YOLO model for object detection"""
        try:
            self.status_var.set("Loading model...")
            self.model = YOLO("best.pt")

            # Try to move model to GPU if available
            try:
                self.model.to("cuda")
                self.status_var.set("Model loaded on GPU")
            except:
                self.status_var.set("Model loaded on CPU")

            self.model_loaded = True
            messagebox.showinfo("Model Loaded", "YOLO model loaded successfully")
        except Exception as e:
            self.status_var.set("Error loading model")
            messagebox.showerror("Error", f"Failed to load model: {str(e)}")
            self.model_loaded = False

    def load_global_count(self):
        """Load global accident count from file if it exists"""
        count_file = "accident_count.txt"
        try:
            if os.path.exists(count_file):
                with open(count_file, 'r') as f:
                    self.global_accident_count = int(f.read().strip())
                    self.total_accidents = self.global_accident_count
                    self.total_accidents_var.set(str(self.total_accidents))
        except Exception as e:
            print(f"Error loading global count: {str(e)}")
            self.global_accident_count = 0

    def save_global_count(self):
        """Save global accident count to file"""
        count_file = "accident_count.txt"
        try:
            with open(count_file, 'w') as f:
                f.write(str(self.global_accident_count))
        except Exception as e:
            print(f"Error saving global count: {str(e)}")

    def select_files(self):
        """Open file dialog to select image or video files"""
        filetypes = (
            ("Media files", "*.jpg *.jpeg *.png *.mp4 *.avi *.mov"),
            ("Image files", "*.jpg *.jpeg *.png"),
            ("Video files", "*.mp4 *.avi *.mov"),
            ("All files", "*.*")
        )

        files = filedialog.askopenfilenames(
            title="Select files",
            filetypes=filetypes
        )

        if files:
            self.file_paths = list(files)
            self.files_listbox.delete(0, tk.END)

            for file in self.file_paths:
                self.files_listbox.insert(tk.END, os.path.basename(file))

            self.current_file_index = 0
            self.update_file_controls()
            self.load_current_file()

    def on_file_select(self, event):
        """Handle selection of a file in the listbox"""
        if self.files_listbox.curselection():
            self.current_file_index = self.files_listbox.curselection()[0]
            self.load_current_file()

    def load_current_file(self):
        """Load and display the currently selected file"""
        if not self.file_paths or self.current_file_index >= len(self.file_paths):
            return

        # Reset state
        self.stop_video_playback()

        file_path = self.file_paths[self.current_file_index]
        file_ext = os.path.splitext(file_path)[1].lower()

        # Check if it's a video or image
        if file_ext in ['.mp4', '.avi', '.mov']:
            # Video file
            self.cap = cv2.VideoCapture(file_path)
            if not self.cap.isOpened():
                messagebox.showerror("Error", f"Could not open video: {file_path}")
                return

            # Get first frame to display
            ret, frame = self.cap.read()
            if ret:
                self.current_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                self.display_image(self.current_frame)
                self.play_btn.configure(state=tk.NORMAL)
            else:
                self.play_btn.configure(state=tk.DISABLED)
        else:
            # Image file
            try:
                self.current_frame = cv2.imread(file_path)
                if self.current_frame is None:
                    raise Exception("Failed to load image")

                self.current_frame = cv2.cvtColor(self.current_frame, cv2.COLOR_BGR2RGB)
                self.display_image(self.current_frame)
                self.play_btn.configure(state=tk.DISABLED)
            except Exception as e:
                messagebox.showerror("Error", f"Could not open image: {str(e)}")

        # Update status
        filename = os.path.basename(file_path)
        self.status_var.set(f"Loaded: {filename}")
        self.files_listbox.selection_clear(0, tk.END)
        self.files_listbox.selection_set(self.current_file_index)
        self.files_listbox.see(self.current_file_index)

    def display_image(self, img):
        """Display an image on the canvas"""
        if img is None:
            return

        # Resize image to fit canvas while maintaining aspect ratio
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()

        if canvas_width <= 1 or canvas_height <= 1:
            # Canvas not yet properly sized, retry after a short delay
            self.root.after(100, lambda: self.display_image(img))
            return

        img_height, img_width = img.shape[:2]

        # Calculate scaling factor to fit canvas
        width_ratio = canvas_width / img_width
        height_ratio = canvas_height / img_height
        scale_factor = min(width_ratio, height_ratio)

        new_width = int(img_width * scale_factor)
        new_height = int(img_height * scale_factor)

        # Resize image
        resized_img = cv2.resize(img, (new_width, new_height), interpolation=cv2.INTER_AREA)

        # Convert to PhotoImage
        self.photo = ImageTk.PhotoImage(image=Image.fromarray(resized_img))

        # Clear canvas and display image
        self.canvas.delete("all")
        self.canvas.create_image(
            (canvas_width - new_width) // 2,
            (canvas_height - new_height) // 2,
            anchor=tk.NW,
            image=self.photo
        )

    def toggle_play(self):
        """Toggle video playback start/stop"""
        if not self.cap:
            return

        if self.is_video_playing:
            self.stop_video_playback()
            self.play_btn.configure(text="▶ Play")
        else:
            self.is_video_playing = True
            self.play_btn.configure(text="⏸ Pause")
            self.play_video()

    def play_video(self):
        """Play the current video frame by frame"""
        if not self.is_video_playing or not self.cap:
            return

        ret, frame = self.cap.read()
        if ret:
            self.current_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            self.display_image(self.current_frame)
            # Continue playing after a delay (based on FPS)
            fps = self.cap.get(cv2.CAP_PROP_FPS)
            delay = int(1000 / fps) if fps > 0 else 33  # Default to ~30 FPS
            self.root.after(delay, self.play_video)
        else:
            # End of video
            self.cap.set(cv2.CAP_PROP_POS_FRAMES, 0)  # Reset to beginning
            self.stop_video_playback()

    def stop_video_playback(self):
        """Stop video playback"""
        self.is_video_playing = False
        self.play_btn.configure(text="▶ Play")

    def previous_file(self):
        """Navigate to the previous file"""
        if self.file_paths and self.current_file_index > 0:
            self.current_file_index -= 1
            self.load_current_file()
            self.update_file_controls()

    def next_file(self):
        """Navigate to the next file"""
        if self.file_paths and self.current_file_index < len(self.file_paths) - 1:
            self.current_file_index += 1
            self.load_current_file()
            self.update_file_controls()

    def update_file_controls(self):
        """Update navigation buttons based on current file index"""
        if not self.file_paths:
            self.prev_btn.configure(state=tk.DISABLED)
            self.next_btn.configure(state=tk.DISABLED)
            return

        self.prev_btn.configure(state=tk.NORMAL if self.current_file_index > 0 else tk.DISABLED)
        self.next_btn.configure(state=tk.NORMAL if self.current_file_index < len(self.file_paths) - 1 else tk.DISABLED)

    def process_selected(self):
        """Process the currently selected file for accident detection"""
        if not self.model_loaded:
            messagebox.showerror("Error", "Model not loaded. Please load the model first.")
            return

        if not self.file_paths:
            messagebox.showinfo("Info", "No files selected. Please select files to process.")
            return

        # Update parameters from UI
        self.confidence_threshold = self.confidence_var.get()
        self.detection_time_threshold = self.persistence_var.get()

        # Disable controls during processing
        self.process_btn.configure(state=tk.DISABLED)
        self.stop_btn.configure(state=tk.NORMAL)

        # Don't reset total accidents, only current file accidents
        # self.total_accidents = 0  # Remove this line to maintain count

        # Start processing in a separate thread
        self.stop_processing = False
        self.processing_thread = threading.Thread(target=self.process_files)
        self.processing_thread.daemon = True
        self.processing_thread.start()

    def process_files(self):
        """Process files for accident detection"""
        # Excel file setup for logging accident data
        excel_file = "accident_log.xlsx"
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Accident Log"
            ws.append(["Date", "File", "Type", "File Accident Count", "Global Count"])
            wb.save(excel_file)

        # Initialize total accident count (don't reset between files)
        # self.total_accidents = 0  # This line is removed to maintain count

        for index, file_path in enumerate(self.file_paths):
            if self.stop_processing:
                break

            # Update UI to show current file
            self.root.after(0, lambda idx=index: self.update_current_file(idx))

            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_path)[1].lower()

            self.root.after(0, lambda msg=f"Processing: {file_name}": self.status_var.set(msg))

            # Reset accident count for current file only, not total
            self.accident_count = 0
            self.accident_locations = []
            self.root.after(0, lambda: self.current_accidents_var.set("0"))

            if file_ext.lower() in ['.mp4', '.avi', '.mov']:
                # Process video
                self.process_video(file_path, file_name)
            else:
                # Process image
                self.process_image(file_path, file_name)

        # Processing complete
        if not self.stop_processing:
            self.root.after(0, lambda: self.status_var.set("Processing complete"))
            self.root.after(0, lambda: messagebox.showinfo("Complete",
                                                           f"Processing complete. Total accidents detected: {self.total_accidents}"))
        else:
            self.root.after(0, lambda: self.status_var.set("Processing stopped"))

        # Re-enable controls
        self.root.after(0, lambda: self.process_btn.configure(state=tk.NORMAL))
        self.root.after(0, lambda: self.stop_btn.configure(state=tk.DISABLED))

    def update_current_file(self, index):
        """Update UI to show the file being processed"""
        self.current_file_index = index
        self.files_listbox.selection_clear(0, tk.END)
        self.files_listbox.selection_set(index)
        self.files_listbox.see(index)
        self.load_current_file()

    def process_video(self, video_path, video_name):
        """Process a video file for accident detection"""
        # Initialize video capture
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            self.root.after(0, lambda: messagebox.showerror("Error", f"Unable to open video {video_path}"))
            return

        # Get video properties
        fps = cap.get(cv2.CAP_PROP_FPS)

        # Initialize variables for detection and tracking
        accident_count = 0
        accident_counted = False
        tracker = None
        tracking = False
        accumulated_detection_time = 0.0
        persistent_display_time = 0.0

        # Time tracking for real-time accident detection
        prev_time = time.time()

        frame_count = 0

        while True:
            if self.stop_processing:
                break

            ret, img = cap.read()
            if not ret:
                break  # End of video

            frame_count += 1

            # Process every 3rd frame to speed up processing (can be adjusted)
            if frame_count % 3 != 0:
                continue

            # Capture current time for timing
            current_time = time.time()
            dt = current_time - prev_time
            prev_time = current_time

            # Process frame for accident detection
            if not tracking:
                # Object detection phase
                results = self.model(img, stream=True, batch=self.batch_size)
                candidate_box = None
                candidate_conf = 0.0

                # Process detection results
                for r in results:
                    boxes = r.boxes
                    for box in boxes:
                        x1, y1, x2, y2 = box.xyxy[0]
                        x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                        w, h = x2 - x1, y2 - y1
                        conf = float(box.conf[0])
                        cls = int(box.cls[0])

                        # Draw bounding box and class label
                        cvzone.cornerRect(img, (x1, y1, w, h))
                        cvzone.putTextRect(img, f"{self.classNames[cls]} {conf:.2f}",
                                           (max(0, x1), max(35, y1)),
                                           scale=1, thickness=1)

                        # Select candidate with highest confidence
                        if conf >= self.confidence_threshold and conf > candidate_conf:
                            candidate_box = (x1, y1, w, h)
                            candidate_conf = conf

                # Initialize tracking if a candidate is detected consistently
                if candidate_box is not None:
                    accumulated_detection_time += dt
                    if accumulated_detection_time >= self.detection_time_threshold:
                        try:
                            tracker = cv2.TrackerKCF_create()
                        except AttributeError:
                            try:
                                tracker = cv2.legacy.TrackerKCF_create()
                            except Exception:
                                tracker = cv2.TrackerCSRT_create()  # Fallback

                        tracker.init(img, candidate_box)
                        tracking = True
                        persistent_display_time = 0.0

                        # Log accident once per detected event
                        if not accident_counted:
                            accident_count += 1
                            self.accident_count += 1
                            accident_counted = True

                            # Add to persistent accident locations
                            x, y, w, h = candidate_box
                            self.accident_locations.append({
                                'box': candidate_box,
                                'time': time.time(),
                                'number': self.global_accident_count + 1
                                # Use global count + 1 (will be incremented in log)
                            })

                            # Log to Excel (will increment global counter)
                            self.log_accident_to_excel("Video", video_name)

                            # Update UI
                            self.root.after(0, lambda c=str(self.accident_count): self.current_accidents_var.set(c))
                            self.root.after(0, lambda t=str(self.total_accidents): self.total_accidents_var.set(t))
                else:
                    accumulated_detection_time = 0.0
            else:
                # Tracking phase
                success_track, bbox = tracker.update(img)
                if success_track:
                    x, y, w, h = [int(v) for v in bbox]
                    cvzone.cornerRect(img, (x, y, w, h))

                    # Update persistent display time
                    persistent_display_time += dt

                    # Display "Accident is persistent" with duration
                    global_count = self.accident_locations[-1]['number'] if self.accident_locations else 0
                    persistence_text = f"Accident #{global_count} - persistent ({persistent_display_time:.1f}s)"
                    cvzone.putTextRect(img, persistence_text,
                                       (max(0, x), max(35, y)),
                                       scale=1, thickness=1,
                                       colorR=(0, 0, 255))  # Red for persistent accidents
                else:
                    tracking = False
                    tracker = None
                    accumulated_detection_time = 0.0

                    # Allow a new accident to be counted after tracking is lost
                    accident_counted = False

            # Display all previously detected accidents
            for i, accident in enumerate(self.accident_locations):
                x, y, w, h = accident['box']
                accident_number = accident['number']
                accident_time = time.strftime("%H:%M:%S", time.localtime(accident['time']))

                # Draw a small marker at the accident location
                cv2.circle(img, (x + w // 2, y + h // 2), 5, (0, 0, 255), -1)

                # Only show detailed text for the most recent accidents to avoid cluttering
                if i >= max(0, len(self.accident_locations) - 3):
                    marker_text = f"Accident #{accident_number} at {accident_time}"
                    cv2.putText(img, marker_text, (x + w // 2 + 10, y + h // 2),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

            # Add information overlay with modern styling
            self.add_info_overlay(img)

            # Display the frame
            display_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            self.root.after(0, lambda img=display_img: self.display_image(img))

            # Control processing speed to not overwhelm the UI
            time.sleep(0.01)

        # Release video capture
        cap.release()

    def process_image(self, image_path, image_name):
        """Process a single image for accident detection"""
        # Load image
        img = cv2.imread(image_path)
        if img is None:
            self.root.after(0, lambda: messagebox.showerror("Error", f"Unable to open image {image_path}"))
            return

        # Process image for accident detection
        results = self.model(img, stream=True, batch=self.batch_size)

        has_accident = False

        # Process detection results
        for r in results:
            boxes = r.boxes
            for box in boxes:
                x1, y1, x2, y2 = box.xyxy[0]
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                w, h = x2 - x1, y2 - y1
                conf = float(box.conf[0])
                cls = int(box.cls[0])

                if conf >= self.confidence_threshold:
                    # Draw bounding box and class label
                    cvzone.cornerRect(img, (x1, y1, w, h))
                    cvzone.putTextRect(img, f"{self.classNames[cls]} {conf:.2f}",
                                       (max(0, x1), max(35, y1)),
                                       scale=1, thickness=1)

                    # Mark as accident
                    has_accident = True

                    # Add to accident locations
                    self.accident_locations.append({
                        'box': (x1, y1, w, h),
                        'time': time.time(),
                        'number': self.global_accident_count + 1  # Will be incremented in log
                    })

        if has_accident:
            self.accident_count += 1

            # Log to Excel (will increment global counter)
            self.log_accident_to_excel("Image", image_name)

            # Update UI
            self.root.after(0, lambda c=str(self.accident_count): self.current_accidents_var.set(c))
            self.root.after(0, lambda t=str(self.total_accidents): self.total_accidents_var.set(t))

        # Display overlay text
        self.add_info_overlay(img)

        # Display the image
        display_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.root.after(0, lambda img=display_img: self.display_image(img))

        # Small delay to allow UI to update
        time.sleep(0.5)

    def add_info_overlay(self, img):
        """Add stylized information overlay to the image"""
        # Create a semi-transparent overlay at the bottom
        h, w = img.shape[:2]
        overlay = img.copy()
        footer_height = 80
        cv2.rectangle(overlay, (0, h - footer_height), (w, h), (25, 25, 25), -1)
        cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

        # Add file information
        if self.file_paths and self.current_file_index < len(self.file_paths):
            file_name = os.path.basename(self.file_paths[self.current_file_index])
            cv2.putText(img, f"File: {file_name}", (10, h - footer_height + 20),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

        # Add detection parameters
        params_text = f"Conf: {self.confidence_threshold:.2f} | Persist: {self.detection_time_threshold:.1f}s"
        cv2.putText(img, params_text, (10, h - footer_height + 45),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

        # Add accident count in a prominent style
        count_text = f"File: {self.accident_count} | Total: {self.global_accident_count}"
        text_size = cv2.getTextSize(count_text, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)[0]
        cv2.putText(img, count_text, (w - text_size[0] - 20, h - footer_height + 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (50, 205, 255), 2)

    def log_accident_to_excel(self, file_type, file_name):
        """Log accident data to Excel file with global counter"""
        try:
            excel_file = "accident_log.xlsx"
            if not os.path.exists(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Accident Log"
                ws.append(["Date", "File", "Type", "File Accident Count", "Global Count"])
                wb.save(excel_file)

            wb = load_workbook(excel_file)
            ws = wb.active
            current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            self.global_accident_count += 1  # Increment global count
            ws.append([current_date, file_name, file_type, self.accident_count, self.global_accident_count])
            wb.save(excel_file)

            # Save global count to separate file
            self.save_global_count()

            # Update total accidents with global count
            self.total_accidents = self.global_accident_count

        except Exception as e:
            self.root.after(0, lambda: self.status_var.set(f"Error logging to Excel: {str(e)}"))

    def stop_processing_cmd(self):
        """Stop ongoing processing"""
        self.stop_processing = True
        self.status_var.set("Stopping processing...")

    def view_log(self):
        """View the accident log Excel file"""
        excel_file = "accident_log.xlsx"
        if not os.path.exists(excel_file):
            messagebox.showinfo("Info", "No log file exists yet.")
            return

        # Try to open Excel file with default application
        try:
            if os.name == 'nt':  # Windows
                os.startfile(excel_file)
            elif os.name == 'darwin':  # macOS
                os.system(f"open {excel_file}")
            else:  # Linux
                os.system(f"xdg-open {excel_file}")
        except Exception as e:
            messagebox.showerror("Error", f"Could not open log file: {str(e)}")

    def show_about(self):
        about_text = """
        Accident Detection System

        A modern GUI application for detecting accidents in images and videos
        using advanced YOLO object detection.

        Features:
        - Process both images and videos
        - Adjustable detection parameters
        - Real-time tracking of accidents
        - Detailed logging of detected accidents
        - Modern, easy-to-use interface

        Version: 1.0
        """
        messagebox.showinfo("About", about_text)


def run_gui():
    """Run the main GUI application"""
    root = tk.Tk()
    # Set app icon
    try:
        # If you have an icon file:
        # root.iconbitmap("app_icon.ico")  # for Windows
        # root.iconphoto(True, tk.PhotoImage(file="app_icon.png"))  # for Linux/Mac
        pass
    except:
        pass

    # Set theme - using ttk themed widgets
    style = ttk.Style()
    if os.name == 'nt':  # Windows
        style.theme_use('vista')
    elif os.name == 'darwin':  # macOS
        style.theme_use('aqua')
    else:  # Linux and others
        style.theme_use('clam')

    # Configure some styles
    style.configure('TButton', font=('Arial', 10))
    style.configure('TLabel', font=('Arial', 10))
    style.configure('TLabelframe.Label', font=('Arial', 10, 'bold'))

    app = AccidentDetectionGUI(root)
    root.mainloop()


def process_files_cli(file_paths, confidence_threshold=0.7, detection_time_threshold=2.0, batch_size=1):
    """CLI version of the accident detection system for backward compatibility"""
    # Create Excel log file if it doesn't exist
    excel_file = "accident_log.xlsx"
    if not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Accident Log"
        ws.append(["Date", "File", "Type", "Accident Count", "Running Total"])
        wb.save(excel_file)

    # Initialize YOLO model
    model = YOLO("best.pt")
    try:
        model.to("cuda")  # Move model to GPU if available
    except:
        print("CUDA not available, using CPU")

    classNames = ['moderate', 'severe']
    total_accidents = 0

    # Load global accident count
    count_file = "accident_count.txt"
    try:
        if os.path.exists(count_file):
            with open(count_file, 'r') as f:
                global_count = int(f.read().strip())
                total_accidents = global_count
    except Exception as e:
        print(f"Error loading global count: {str(e)}")
        global_count = 0

    for file_path in file_paths:
        file_name = os.path.basename(file_path)
        file_ext = os.path.splitext(file_path)[1].lower()
        print(f"Processing: {file_name}")

        # Reset current file accident count but maintain total
        file_accident_count = 0

        if file_ext in ['.mp4', '.avi', '.mov']:
            # Process video file
            cap = cv2.VideoCapture(file_path)
            if not cap.isOpened():
                print(f"Error: Unable to open video {file_path}")
                continue

            # Get video properties
            fps = cap.get(cv2.CAP_PROP_FPS)
            frame_interval = int(1000 / fps)

            # Initialize variables for detection and tracking
            accident_counted = False
            tracker = None
            tracking = False
            accumulated_detection_time = 0.0
            persistent_display_time = 0.0
            accident_locations = []

            # Time tracking
            prev_time = time.time()

            while True:
                ret, img = cap.read()
                if not ret:
                    break

                # Capture timing
                current_time = time.time()
                dt = current_time - prev_time
                prev_time = current_time

                if not tracking:
                    # Detection phase
                    results = model(img, stream=True, batch=batch_size)
                    candidate_box = None
                    candidate_conf = 0.0

                    for r in results:
                        boxes = r.boxes
                        for box in boxes:
                            x1, y1, x2, y2 = box.xyxy[0]
                            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                            w, h = x2 - x1, y2 - y1
                            conf = float(box.conf[0])
                            cls = int(box.cls[0])

                            # Draw bounding box
                            cvzone.cornerRect(img, (x1, y1, w, h))
                            cvzone.putTextRect(img, f"{classNames[cls]} {conf:.2f}",
                                               (max(0, x1), max(35, y1)),
                                               scale=1, thickness=1)

                            # Select highest confidence detection
                            if conf >= confidence_threshold and conf > candidate_conf:
                                candidate_box = (x1, y1, w, h)
                                candidate_conf = conf

                    if candidate_box is not None:
                        accumulated_detection_time += dt
                        if accumulated_detection_time >= detection_time_threshold:
                            try:
                                tracker = cv2.TrackerKCF_create()
                            except AttributeError:
                                try:
                                    tracker = cv2.legacy.TrackerKCF_create()
                                except Exception:
                                    tracker = cv2.TrackerCSRT_create()

                            tracker.init(img, candidate_box)
                            tracking = True
                            persistent_display_time = 0.0

                            # Log accident
                            if not accident_counted:
                                file_accident_count += 1
                                total_accidents += 1
                                accident_counted = True

                                # Add to accident locations
                                x, y, w, h = candidate_box
                                accident_locations.append({
                                    'box': candidate_box,
                                    'time': time.time(),
                                    'number': total_accidents  # Use total count for global numbering
                                })

                                # Log to Excel
                                wb = load_workbook(excel_file)
                                ws = wb.active
                                current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                ws.append([current_date, file_name, "Video", file_accident_count, total_accidents])
                                wb.save(excel_file)

                                # Update global count file
                                with open(count_file, 'w') as f:
                                    f.write(str(total_accidents))

                                print(
                                    f"Accident #{total_accidents} detected in {file_name} (File count: {file_accident_count})")
                    else:
                        accumulated_detection_time = 0.0
                else:
                    # Tracking phase
                    success_track, bbox = tracker.update(img)
                    if success_track:
                        x, y, w, h = [int(v) for v in bbox]
                        cvzone.cornerRect(img, (x, y, w, h))

                        # Update persistence time
                        persistent_display_time += dt

                        # Display persistent message
                        persistence_text = f"Accident #{total_accidents} - persistent ({persistent_display_time:.1f}s)"
                        cvzone.putTextRect(img, persistence_text,
                                           (max(0, x), max(35, y)),
                                           scale=1, thickness=1,
                                           colorR=(0, 0, 255))
                    else:
                        tracking = False
                        tracker = None
                        accumulated_detection_time = 0.0
                        accident_counted = False

                # Display all detected accidents
                for i, accident in enumerate(accident_locations):
                    x, y, w, h = accident['box']
                    accident_number = accident['number']
                    accident_time = time.strftime("%H:%M:%S", time.localtime(accident['time']))

                    # Draw marker
                    cv2.circle(img, (x + w // 2, y + h // 2), 5, (0, 0, 255), -1)

                    # Show text for recent accidents
                    if i >= max(0, len(accident_locations) - 5):
                        marker_text = f"Accident #{accident_number} at {accident_time}"
                        cv2.putText(img, marker_text, (x + w // 2 + 10, y + h // 2),
                                    cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)

                # Add information overlay
                h, w = img.shape[:2]
                overlay = img.copy()
                cv2.rectangle(overlay, (0, h - 80), (w, h), (25, 25, 25), -1)
                cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

                # Display video info
                cv2.putText(img, f"Video: {file_name}", (10, h - 55),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

                # Display parameters
                params_text = f"Conf: {confidence_threshold:.2f} | Persist: {detection_time_threshold:.1f}s"
                cv2.putText(img, params_text, (10, h - 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

                # Display accident counts
                count_text = f"File: {file_accident_count} | Total: {total_accidents}"
                text_size = cv2.getTextSize(count_text, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)[0]
                cv2.putText(img, count_text, (w - text_size[0] - 20, h - 40),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.8, (50, 205, 255), 2)

                # Show image
                cv2.imshow("Accident Detection", img)

                # Maintain original video speed
                elapsed_ms = int(dt * 1000)
                wait_ms = max(1, frame_interval - elapsed_ms)

                # Wait for key press
                key = cv2.waitKey(wait_ms)
                if key == ord('q'):
                    break
                elif key == ord('p'):
                    print("Video paused. Press any key to continue.")
                    cv2.waitKey(0)

            # Release video
            cap.release()

        else:
            # Process image file
            img = cv2.imread(file_path)
            if img is None:
                print(f"Error: Unable to open image {file_path}")
                continue

            # Object detection
            results = model(img, stream=True, batch=batch_size)

            accident_detected = False

            for r in results:
                boxes = r.boxes
                for box in boxes:
                    x1, y1, x2, y2 = box.xyxy[0]
                    x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
                    w, h = x2 - x1, y2 - y1
                    conf = float(box.conf[0])
                    cls = int(box.cls[0])

                    if conf >= confidence_threshold:
                        # Draw bounding box
                        cvzone.cornerRect(img, (x1, y1, w, h))
                        cvzone.putTextRect(img, f"{classNames[cls]} {conf:.2f}",
                                           (max(0, x1), max(35, y1)),
                                           scale=1, thickness=1)

                        accident_detected = True

            if accident_detected:
                file_accident_count = 1
                total_accidents += 1

                # Log to Excel
                wb = load_workbook(excel_file)
                ws = wb.active
                current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ws.append([current_date, file_name, "Image", file_accident_count, total_accidents])
                wb.save(excel_file)

                # Update global count file
                with open(count_file, 'w') as f:
                    f.write(str(total_accidents))

                print(f"Accident #{total_accidents} detected in image: {file_name}")

            # Display information
            h, w = img.shape[:2]
            overlay = img.copy()
            cv2.rectangle(overlay, (0, h - 80), (w, h), (25, 25, 25), -1)
            cv2.addWeighted(overlay, 0.7, img, 0.3, 0, img)

            # Display image info
            cv2.putText(img, f"Image: {file_name}", (10, h - 55),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

            # Display parameters
            params_text = f"Confidence: {confidence_threshold:.2f}"
            cv2.putText(img, params_text, (10, h - 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.6, (200, 200, 200), 1)

            # Display accident info with total count
            if accident_detected:
                result_text = f"Accident #{total_accidents} Detected"
            else:
                result_text = "No Accident"
            text_color = (50, 205, 255) if accident_detected else (200, 200, 200)
            text_size = cv2.getTextSize(result_text, cv2.FONT_HERSHEY_SIMPLEX, 0.8, 2)[0]
            cv2.putText(img, result_text, (w - text_size[0] - 20, h - 40),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, text_color, 2)

            # Show image
            cv2.imshow("Accident Detection", img)
            cv2.waitKey(0)

    # Cleanup
    cv2.destroyAllWindows()
    print(f"Total accidents detected across all files: {total_accidents}")
    return total_accidents


if __name__ == "__main__":
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Accident Detection System')
    parser.add_argument('--gui', action='store_true', help='Run with GUI interface')
    parser.add_argument('--files', nargs='+', type=str, default=[],
                        help='List of file paths (images or videos) to process')
    parser.add_argument('--confidence', type=float, default=0.7,
                        help='Confidence threshold for detection (default: 0.7)')
    parser.add_argument('--persistence', type=float, default=2.0,
                        help='Time in seconds to confirm detection (default: 2.0)')
    args = parser.parse_args()

    # Add your specific files as default if no files provided
    if not args.files:
        # Default image and video paths
        args.files = [
            r"C:\Users\jjnow\OneDrive\Desktop\repositories\IEEE_safety_system_roads\Accident.png",
            r"C:\Users\jjnow\OneDrive\Desktop\repositories\IEEE_safety_system_roads\p050q1w8.jpg",
            r"C:\Users\jjnow\OneDrive\Desktop\repositories\IEEE_safety_system_roads\video_1.mp4"
        ]

    # Determine whether to run in GUI or CLI mode
    if args.gui:
        print("Starting GUI mode...")
        run_gui()
    else:
        # Process files in CLI mode
        print("Starting CLI mode...")
        process_files_cli(
            args.files,
            confidence_threshold=args.confidence,
            detection_time_threshold=args.persistence
        )