import cv2
import cvzone
import time
import os
import numpy as np
from ultralytics import YOLO
from openpyxl import Workbook, load_workbook
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import threading
from pathlib import Path


class ModernAccidentGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("AI Crash Detection")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 750)
        self.root.configure(bg="#0a0a0f")

        self.BG = "#0a0a0f"
        self.PANEL = "#12121a"
        self.CARD = "#1a1a2e"
        self.BORDER = "#2a2a3e"
        self.TEXT = "#e0e0e8"
        self.DIM = "#6a6a7e"
        self.ACCENT = "#00d4ff"
        self.RED = "#ff3a5c"
        self.RED_DIM = "#aa1a3c"
        self.GREEN = "#00e88f"
        self.ORANGE = "#ff8a2e"
        self.YELLOW = "#ffd32e"

        self.model = None
        self.model_loaded = False
        self.classNames = ['moderate', 'severe']
        self.confidence_threshold = 0.7
        self.detection_time_threshold = 2.0

        self.file_paths = []
        self.current_file_index = 0
        self.is_processing = False
        self.cap = None
        self.current_frame = None
        self.accident_count = 0
        self.total_accidents = 0
        self.global_accident_count = 0
        self.stop_processing = False
        self.processing_thread = None
        self.severity_label = ""
        self.current_confidence = 0.0
        self.fps_display = 0.0
        self.frame_count_display = 0
        self.total_frames = 0

        self.build_ui()
        self.load_model_async()
        self.load_global_count()

    def build_ui(self):
        self.root.grid_columnconfigure(1, weight=1)
        self.root.grid_rowconfigure(0, weight=1)

        sidebar = tk.Frame(self.root, bg=self.PANEL, width=300)
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar.grid_propagate(False)
        self.build_sidebar(sidebar)

        main = tk.Frame(self.root, bg=self.BG)
        main.grid(row=0, column=1, sticky="nsew")
        main.grid_rowconfigure(0, weight=1)
        main.grid_columnconfigure(0, weight=1)
        self.build_main_view(main)

        bottom = tk.Frame(self.root, bg=self.PANEL, height=50)
        bottom.grid(row=1, column=0, columnspan=2, sticky="ew")
        bottom.grid_propagate(False)
        self.build_status_bar(bottom)

    def build_sidebar(self, parent):
        header = tk.Frame(parent, bg=self.CARD, height=80)
        header.pack(fill="x", padx=12, pady=(12, 0))
        header.pack_propagate(False)

        title_frame = tk.Frame(header, bg=self.CARD)
        title_frame.pack(expand=True)

        dot = tk.Label(title_frame, text="●", fg=self.RED, bg=self.CARD, font=("Consolas", 14))
        dot.pack(side="left", padx=(0, 6))
        self.pulse_dot = dot
        self.pulse_state = False

        tk.Label(title_frame, text="CRASH DETECT", fg=self.TEXT, bg=self.CARD,
                 font=("Consolas", 16, "bold")).pack(side="left")

        tk.Label(header, text="AI-POWERED SAFETY SYSTEM", fg=self.DIM, bg=self.CARD,
                 font=("Consolas", 8)).pack()

        self.model_status = tk.Label(parent, text="⏳ Loading model...", fg=self.ORANGE, bg=self.PANEL,
                                     font=("Consolas", 9))
        self.model_status.pack(fill="x", padx=16, pady=(10, 0))

        self.build_section(parent, "INPUT")
        btn_frame = tk.Frame(parent, bg=self.PANEL)
        btn_frame.pack(fill="x", padx=16, pady=(4, 0))

        self.select_btn = self.make_button(btn_frame, "SELECT FILES", self.select_files, self.ACCENT)
        self.select_btn.pack(fill="x", pady=(0, 6))

        self.file_list_frame = tk.Frame(parent, bg=self.CARD, highlightbackground=self.BORDER,
                                        highlightthickness=1, height=120)
        self.file_list_frame.pack(fill="x", padx=16, pady=(0, 4))
        self.file_list_frame.pack_propagate(False)

        self.file_list_label = tk.Label(self.file_list_frame, text="No files selected", fg=self.DIM,
                                        bg=self.CARD, font=("Consolas", 9), anchor="nw", justify="left")
        self.file_list_label.pack(fill="both", expand=True, padx=8, pady=6)

        self.build_section(parent, "PARAMETERS")

        param_frame = tk.Frame(parent, bg=self.PANEL)
        param_frame.pack(fill="x", padx=16, pady=(4, 0))

        tk.Label(param_frame, text="CONFIDENCE", fg=self.DIM, bg=self.PANEL,
                 font=("Consolas", 8)).pack(anchor="w")
        self.conf_scale = tk.Scale(param_frame, from_=0.1, to=1.0, resolution=0.05, orient="horizontal",
                                   bg=self.PANEL, fg=self.TEXT, troughcolor=self.CARD,
                                   highlightthickness=0, sliderrelief="flat", bd=0,
                                   activebackground=self.ACCENT, font=("Consolas", 9), length=200)
        self.conf_scale.set(0.7)
        self.conf_scale.pack(fill="x")

        tk.Label(param_frame, text="PERSISTENCE (s)", fg=self.DIM, bg=self.PANEL,
                 font=("Consolas", 8)).pack(anchor="w", pady=(6, 0))
        self.pers_scale = tk.Scale(param_frame, from_=0.5, to=5.0, resolution=0.5, orient="horizontal",
                                   bg=self.PANEL, fg=self.TEXT, troughcolor=self.CARD,
                                   highlightthickness=0, sliderrelief="flat", bd=0,
                                   activebackground=self.ACCENT, font=("Consolas", 9), length=200)
        self.pers_scale.set(2.0)
        self.pers_scale.pack(fill="x")

        self.build_section(parent, "ACTIONS")

        action_frame = tk.Frame(parent, bg=self.PANEL)
        action_frame.pack(fill="x", padx=16, pady=(4, 0))

        self.process_btn = self.make_button(action_frame, "▶  RUN DETECTION", self.process_selected, self.GREEN)
        self.process_btn.pack(fill="x", pady=(0, 6))

        self.stop_btn = self.make_button(action_frame, "■  STOP", self.stop_processing_cmd, self.RED)
        self.stop_btn.pack(fill="x")
        self.stop_btn.configure(state="disabled")

        self.build_section(parent, "STATISTICS")

        stats_frame = tk.Frame(parent, bg=self.CARD, highlightbackground=self.BORDER, highlightthickness=1)
        stats_frame.pack(fill="x", padx=16, pady=(4, 8))

        self.stat_current = self.make_stat_row(stats_frame, "CURRENT FILE", "0", self.ACCENT)
        self.stat_total = self.make_stat_row(stats_frame, "TOTAL DETECTED", "0", self.RED)
        self.stat_conf = self.make_stat_row(stats_frame, "LAST CONFIDENCE", "—", self.YELLOW)

    def build_main_view(self, parent):
        self.video_frame = tk.Frame(parent, bg="#000000")
        self.video_frame.pack(fill="both", expand=True, padx=12, pady=12)

        self.canvas = tk.Canvas(self.video_frame, bg="#050508", highlightthickness=0, cursor="crosshair")
        self.canvas.pack(fill="both", expand=True)

        self.overlay_frame = tk.Frame(self.video_frame, bg="#050508")
        self.overlay_frame.place(relx=0.5, rely=0.5, anchor="center")

        self.placeholder_label = tk.Label(self.overlay_frame, text="SELECT FILES TO BEGIN",
                                          fg=self.DIM, bg="#050508", font=("Consolas", 14))
        self.placeholder_label.pack()
        self.placeholder_sub = tk.Label(self.overlay_frame, text="Supports MP4, AVI, MOV, JPG, PNG",
                                        fg="#3a3a4e", bg="#050508", font=("Consolas", 9))
        self.placeholder_sub.pack(pady=(4, 0))

        self.alert_frame = tk.Frame(parent, bg=self.BG, height=0)
        self.alert_frame.pack(fill="x", padx=12, pady=(0, 12))
        self.alert_frame.pack_propagate(False)

        self.alert_inner = tk.Frame(self.alert_frame, bg=self.RED_DIM)
        self.alert_inner.pack(fill="both", expand=True)

        self.alert_icon = tk.Label(self.alert_inner, text="⚠", fg="#ffffff", bg=self.RED_DIM,
                                   font=("Consolas", 20, "bold"))
        self.alert_icon.pack(side="left", padx=(16, 8))

        alert_text_frame = tk.Frame(self.alert_inner, bg=self.RED_DIM)
        alert_text_frame.pack(side="left", fill="both", expand=True, pady=8)

        self.alert_title = tk.Label(alert_text_frame, text="CRASH DETECTED", fg="#ffffff",
                                    bg=self.RED_DIM, font=("Consolas", 14, "bold"), anchor="w")
        self.alert_title.pack(fill="x")

        self.alert_detail = tk.Label(alert_text_frame, text="", fg="#ffcccc",
                                     bg=self.RED_DIM, font=("Consolas", 9), anchor="w")
        self.alert_detail.pack(fill="x")

        self.alert_conf = tk.Label(self.alert_inner, text="", fg="#ffffff",
                                   bg=self.RED_DIM, font=("Consolas", 22, "bold"))
        self.alert_conf.pack(side="right", padx=16)

    def build_status_bar(self, parent):
        self.status_label = tk.Label(parent, text="Ready", fg=self.DIM, bg=self.PANEL,
                                     font=("Consolas", 9), anchor="w")
        self.status_label.pack(side="left", padx=16, fill="y")

        self.fps_label = tk.Label(parent, text="", fg=self.DIM, bg=self.PANEL,
                                  font=("Consolas", 9), anchor="e")
        self.fps_label.pack(side="right", padx=16, fill="y")

        self.progress_label = tk.Label(parent, text="", fg=self.ACCENT, bg=self.PANEL,
                                       font=("Consolas", 9))
        self.progress_label.pack(side="right", padx=(0, 16), fill="y")

    def build_section(self, parent, title):
        sep = tk.Frame(parent, bg=self.BORDER, height=1)
        sep.pack(fill="x", padx=16, pady=(12, 0))
        tk.Label(parent, text=title, fg=self.DIM, bg=self.PANEL,
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=16, pady=(6, 0))

    def make_button(self, parent, text, command, color):
        fg = "#000000" if color in [self.GREEN, self.YELLOW, self.ACCENT] else "#ffffff"
        btn = tk.Button(parent, text=text, command=command,
                        bg=color, fg=fg, activebackground=color, activeforeground=fg,
                        font=("Consolas", 10, "bold"), relief="flat", bd=0,
                        padx=16, pady=8, cursor="hand2")
        btn.bind("<Enter>", lambda e, b=btn, c=color: b.configure(bg=self._lighten(c)))
        btn.bind("<Leave>", lambda e, b=btn, c=color: b.configure(bg=c))
        return btn

    def _lighten(self, hex_color):
        r = min(255, int(hex_color[1:3], 16) + 30)
        g = min(255, int(hex_color[3:5], 16) + 30)
        b = min(255, int(hex_color[5:7], 16) + 30)
        return f"#{r:02x}{g:02x}{b:02x}"

    def make_stat_row(self, parent, label, value, color):
        row = tk.Frame(parent, bg=self.CARD)
        row.pack(fill="x", padx=10, pady=4)
        tk.Label(row, text=label, fg=self.DIM, bg=self.CARD, font=("Consolas", 8)).pack(side="left")
        val = tk.Label(row, text=value, fg=color, bg=self.CARD, font=("Consolas", 14, "bold"))
        val.pack(side="right")
        return val

    def show_alert(self, severity, confidence):
        color = self.RED if severity == "severe" else self.ORANGE
        self.alert_inner.configure(bg=color)
        self.alert_icon.configure(bg=color)
        self.alert_title.configure(bg=color,
                                   text=f"{'SEVERE' if severity == 'severe' else 'MODERATE'} CRASH DETECTED")
        self.alert_detail.configure(bg=color,
                                    text=f"Accident #{self.global_accident_count} • {datetime.now().strftime('%H:%M:%S')}")
        self.alert_conf.configure(bg=color, text=f"{confidence:.0%}")
        self.alert_frame.configure(height=60)
        self.root.after(4000, self.hide_alert)

    def hide_alert(self):
        self.alert_frame.configure(height=0)

    def pulse_indicator(self):
        if self.is_processing:
            self.pulse_state = not self.pulse_state
            self.pulse_dot.configure(fg=self.RED if self.pulse_state else self.RED_DIM)
            self.root.after(500, self.pulse_indicator)
        else:
            self.pulse_dot.configure(fg=self.RED)

    def load_model_async(self):
        def load():
            try:
                self.model = YOLO("best.pt")
                try:
                    self.model.to("cuda")
                    device = "GPU"
                except:
                    device = "CPU"
                self.model_loaded = True
                self.root.after(0, lambda: self.model_status.configure(
                    text=f"✓ Model ready ({device})", fg=self.GREEN))
            except Exception as e:
                self.root.after(0, lambda: self.model_status.configure(
                    text=f"✗ Model error: {str(e)[:30]}", fg=self.RED))
        threading.Thread(target=load, daemon=True).start()

    def load_global_count(self):
        try:
            f = Path("accident_count.txt")
            if f.exists():
                self.global_accident_count = int(f.read_text().strip())
                self.total_accidents = self.global_accident_count
                self.stat_total.configure(text=str(self.total_accidents))
        except:
            self.global_accident_count = 0

    def save_global_count(self):
        try:
            Path("accident_count.txt").write_text(str(self.global_accident_count))
        except:
            pass

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select media files",
            filetypes=[("Media", "*.jpg *.jpeg *.png *.mp4 *.avi *.mov"), ("All", "*.*")]
        )
        if files:
            self.file_paths = list(files)
            names = [os.path.basename(f) for f in self.file_paths]
            display = "\n".join(f"  {i+1}. {n}" for i, n in enumerate(names))
            self.file_list_label.configure(text=display, fg=self.TEXT)
            self.placeholder_label.configure(text="READY TO PROCESS")
            self.placeholder_sub.configure(text=f"{len(files)} file(s) loaded")
            self.status_label.configure(text=f"{len(files)} files selected")

    def process_selected(self):
        if not self.model_loaded:
            messagebox.showerror("Error", "Model not loaded yet")
            return
        if not self.file_paths:
            messagebox.showinfo("Info", "Select files first")
            return

        self.confidence_threshold = self.conf_scale.get()
        self.detection_time_threshold = self.pers_scale.get()
        self.stop_processing = False
        self.is_processing = True
        self.process_btn.configure(state="disabled")
        self.stop_btn.configure(state="normal")
        self.overlay_frame.place_forget()
        self.pulse_indicator()

        self.processing_thread = threading.Thread(target=self.run_detection, daemon=True)
        self.processing_thread.start()

    def stop_processing_cmd(self):
        self.stop_processing = True
        self.is_processing = False
        self.status_label.configure(text="Stopped")
        self.process_btn.configure(state="normal")
        self.stop_btn.configure(state="disabled")

    def run_detection(self):
        excel_file = "accident_log.xlsx"
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Accident Log"
            ws.append(["Date", "File", "Type", "File Count", "Global Count"])
            wb.save(excel_file)

        for idx, file_path in enumerate(self.file_paths):
            if self.stop_processing:
                break

            file_name = os.path.basename(file_path)
            file_ext = os.path.splitext(file_path)[1].lower()
            self.accident_count = 0
            self.root.after(0, lambda: self.stat_current.configure(text="0"))
            self.root.after(0, lambda n=file_name: self.status_label.configure(text=f"Processing: {n}"))

            if file_ext in ['.mp4', '.avi', '.mov']:
                self.process_video(file_path, file_name, excel_file)
            else:
                self.process_image(file_path, file_name, excel_file)

        self.is_processing = False
        self.root.after(0, lambda: self.process_btn.configure(state="normal"))
        self.root.after(0, lambda: self.stop_btn.configure(state="disabled"))
        if not self.stop_processing:
            self.root.after(0, lambda: self.status_label.configure(text="Detection complete"))

    def process_video(self, path, name, excel_file):
        cap = cv2.VideoCapture(path)
        if not cap.isOpened():
            return

        fps = cap.get(cv2.CAP_PROP_FPS)
        self.total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.frame_count_display = 0

        accident_counted = False
        tracker = None
        tracking = False
        accumulated_time = 0.0
        persistent_time = 0.0
        prev_time = time.time()
        frame_skip = 0
        accident_locations = []

        while True:
            if self.stop_processing:
                break

            ret, img = cap.read()
            if not ret:
                break

            self.frame_count_display += 1
            frame_skip += 1
            if frame_skip % 2 != 0:
                continue

            now = time.time()
            dt = now - prev_time
            prev_time = now
            self.fps_display = 1.0 / dt if dt > 0 else 0

            self.root.after(0, lambda f=self.fps_display: self.fps_label.configure(text=f"FPS: {f:.0f}"))
            self.root.after(0, lambda: self.progress_label.configure(
                text=f"Frame {self.frame_count_display}/{self.total_frames}"))

            if not tracking:
                results = self.model(img, stream=True)
                candidate_box = None
                candidate_conf = 0.0
                candidate_cls = 0

                for r in results:
                    for box in r.boxes:
                        x1, y1, x2, y2 = map(int, box.xyxy[0])
                        w, h = x2 - x1, y2 - y1
                        conf = float(box.conf[0])
                        cls = int(box.cls[0])

                        color = (0, 212, 255) if conf < self.confidence_threshold else (255, 58, 92)
                        cv2.rectangle(img, (x1, y1), (x2, y2), color, 2)
                        label = f"{self.classNames[cls]} {conf:.0%}"
                        self.draw_label(img, label, x1, y1, color)

                        if conf >= self.confidence_threshold and conf > candidate_conf:
                            candidate_box = (x1, y1, w, h)
                            candidate_conf = conf
                            candidate_cls = cls

                if candidate_box:
                    accumulated_time += dt
                    pct = min(accumulated_time / self.detection_time_threshold, 1.0)
                    self.draw_progress_ring(img, candidate_box, pct)

                    if accumulated_time >= self.detection_time_threshold:
                        try:
                            tracker = cv2.TrackerKCF_create()
                        except AttributeError:
                            try:
                                tracker = cv2.legacy.TrackerKCF_create()
                            except:
                                tracker = cv2.TrackerCSRT_create()
                        tracker.init(img, candidate_box)
                        tracking = True
                        persistent_time = 0.0

                        if not accident_counted:
                            self.accident_count += 1
                            self.global_accident_count += 1
                            self.total_accidents = self.global_accident_count
                            accident_counted = True
                            self.current_confidence = candidate_conf
                            self.severity_label = self.classNames[candidate_cls]

                            accident_locations.append({
                                'box': candidate_box, 'time': time.time(),
                                'number': self.global_accident_count,
                                'severity': self.classNames[candidate_cls]
                            })

                            self.log_accident(excel_file, "Video", name)

                            sev = self.severity_label
                            cnf = self.current_confidence
                            self.root.after(0, lambda s=sev, c=cnf: self.show_alert(s, c))
                            self.root.after(0, lambda c=str(self.accident_count): self.stat_current.configure(text=c))
                            self.root.after(0, lambda t=str(self.total_accidents): self.stat_total.configure(text=t))
                            self.root.after(0, lambda c=self.current_confidence: self.stat_conf.configure(
                                text=f"{c:.0%}"))
            else:
                ok, bbox = tracker.update(img)
                if ok:
                    x, y, w, h = map(int, bbox)
                    persistent_time += dt
                    cv2.rectangle(img, (x, y), (x + w, y + h), (255, 58, 92), 3)
                    self.draw_tracking_corners(img, x, y, w, h)
                    tag = f"TRACKING #{self.global_accident_count}  {persistent_time:.1f}s"
                    self.draw_label(img, tag, x, y - 10, (255, 58, 92))
                else:
                    tracking = False
                    tracker = None
                    accumulated_time = 0.0
                    accident_counted = False

            for loc in accident_locations:
                bx, by, bw, bh = loc['box']
                cx, cy = bx + bw // 2, by + bh // 2
                cv2.drawMarker(img, (cx, cy), (255, 58, 92), cv2.MARKER_CROSS, 16, 2)

            self.draw_hud(img, name)
            display = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            self.root.after(0, lambda i=display: self.display_frame(i))
            time.sleep(0.005)

        cap.release()

    def process_image(self, path, name, excel_file):
        img = cv2.imread(path)
        if img is None:
            return

        results = self.model(img, stream=True)
        detected = False

        for r in results:
            for box in r.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                w, h = x2 - x1, y2 - y1
                conf = float(box.conf[0])
                cls = int(box.cls[0])

                if conf >= self.confidence_threshold:
                    cv2.rectangle(img, (x1, y1), (x2, y2), (255, 58, 92), 3)
                    self.draw_tracking_corners(img, x1, y1, w, h)
                    label = f"{self.classNames[cls]} {conf:.0%}"
                    self.draw_label(img, label, x1, y1, (255, 58, 92))
                    detected = True
                    self.current_confidence = conf
                    self.severity_label = self.classNames[cls]

        if detected:
            self.accident_count += 1
            self.global_accident_count += 1
            self.total_accidents = self.global_accident_count
            self.log_accident(excel_file, "Image", name)

            sev = self.severity_label
            cnf = self.current_confidence
            self.root.after(0, lambda s=sev, c=cnf: self.show_alert(s, c))
            self.root.after(0, lambda c=str(self.accident_count): self.stat_current.configure(text=c))
            self.root.after(0, lambda t=str(self.total_accidents): self.stat_total.configure(text=t))
            self.root.after(0, lambda c=self.current_confidence: self.stat_conf.configure(text=f"{c:.0%}"))

        self.draw_hud(img, name)
        display = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        self.root.after(0, lambda i=display: self.display_frame(i))
        time.sleep(1.5)

    def draw_label(self, img, text, x, y, color):
        font = cv2.FONT_HERSHEY_SIMPLEX
        scale, thick = 0.55, 1
        (tw, th), _ = cv2.getTextSize(text, font, scale, thick)
        pad = 6
        cv2.rectangle(img, (x, y - th - pad * 2), (x + tw + pad * 2, y), color, -1)
        cv2.putText(img, text, (x + pad, y - pad), font, scale, (255, 255, 255), thick, cv2.LINE_AA)

    def draw_tracking_corners(self, img, x, y, w, h, length=20, thickness=3):
        c = (255, 58, 92)
        pts = [
            ((x, y), (x + length, y)), ((x, y), (x, y + length)),
            ((x + w, y), (x + w - length, y)), ((x + w, y), (x + w, y + length)),
            ((x, y + h), (x + length, y + h)), ((x, y + h), (x, y + h - length)),
            ((x + w, y + h), (x + w - length, y + h)), ((x + w, y + h), (x + w, y + h - length)),
        ]
        for p1, p2 in pts:
            cv2.line(img, p1, p2, c, thickness)

    def draw_progress_ring(self, img, box, pct):
        x, y, w, h = box
        cx, cy = x + w // 2, y - 25
        radius = 12
        angle = int(360 * pct)
        cv2.ellipse(img, (cx, cy), (radius, radius), -90, 0, 360, (40, 40, 60), 2)
        cv2.ellipse(img, (cx, cy), (radius, radius), -90, 0, angle, (0, 212, 255), 2)

    def draw_hud(self, img, filename):
        h, w = img.shape[:2]
        overlay = img.copy()

        cv2.rectangle(overlay, (0, 0), (w, 36), (10, 10, 15), -1)
        cv2.addWeighted(overlay, 0.8, img, 0.2, 0, img)

        font = cv2.FONT_HERSHEY_SIMPLEX
        cv2.putText(img, "AI CRASH DETECTION", (10, 24), font, 0.55, (0, 212, 255), 1, cv2.LINE_AA)
        ts = datetime.now().strftime("%H:%M:%S")
        cv2.putText(img, ts, (w - 80, 24), font, 0.5, (100, 100, 130), 1, cv2.LINE_AA)

        overlay2 = img.copy()
        cv2.rectangle(overlay2, (0, h - 32), (w, h), (10, 10, 15), -1)
        cv2.addWeighted(overlay2, 0.8, img, 0.2, 0, img)

        cv2.putText(img, filename, (10, h - 10), font, 0.45, (150, 150, 170), 1, cv2.LINE_AA)

        count_text = f"Detections: {self.accident_count}  |  Total: {self.global_accident_count}"
        (tw, _), _ = cv2.getTextSize(count_text, font, 0.45, 1)
        cv2.putText(img, count_text, (w - tw - 10, h - 10), font, 0.45, (255, 58, 92), 1, cv2.LINE_AA)

        if self.accident_count > 0:
            cv2.circle(img, (w - tw - 24, h - 14), 4, (255, 58, 92), -1)

    def display_frame(self, rgb_img):
        if rgb_img is None:
            return

        cw = self.canvas.winfo_width()
        ch = self.canvas.winfo_height()
        if cw <= 1 or ch <= 1:
            self.root.after(50, lambda: self.display_frame(rgb_img))
            return

        ih, iw = rgb_img.shape[:2]
        scale = min(cw / iw, ch / ih)
        nw, nh = int(iw * scale), int(ih * scale)

        resized = cv2.resize(rgb_img, (nw, nh), interpolation=cv2.INTER_AREA)
        self.photo = ImageTk.PhotoImage(image=Image.fromarray(resized))
        self.canvas.delete("all")
        self.canvas.create_image((cw - nw) // 2, (ch - nh) // 2, anchor="nw", image=self.photo)

    def log_accident(self, excel_file, file_type, file_name):
        try:
            if not os.path.exists(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.title = "Accident Log"
                ws.append(["Date", "File", "Type", "File Count", "Global Count"])
                wb.save(excel_file)

            wb = load_workbook(excel_file)
            ws = wb.active
            ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                       file_name, file_type, self.accident_count, self.global_accident_count])
            wb.save(excel_file)
            self.save_global_count()
        except Exception as e:
            self.root.after(0, lambda: self.status_label.configure(text=f"Log error: {str(e)[:40]}"))


def main():
    root = tk.Tk()
    app = ModernAccidentGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
